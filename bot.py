import os
import io
import textwrap
from datetime import datetime

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, BotCommand
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, ConversationHandler, CallbackContext,
    filters, CallbackQueryHandler
)

# === YOUR BOT TOKEN ===
TOKEN = "8444222504:AAHrS8Zs_xTQpQYq80okIboHF6tnw_FfzDc"

# States for ConversationHandler
ENTER_TEXT, CHOOSE_TYPE = range(2)

# Supported types map (accepts user text like 'txt' or '.txt')
SUPPORTED = {
    "txt": "txt",
    "py": "py",
    "css": "css",
    "js": "js",
    "html": "html",
    "json": "json",
    "csv": "csv",
    "xml": "xml",
    "yaml": "yaml",
    "yml": "yaml",
    "pdf": "pdf",
    "xls": "xlsx",   # map xls to xlsx file creation
    "xlsx": "xlsx",
    "md": "md"
}

# ---------- Helpers ----------
def _normalize_ext(ext: str) -> str:
    e = ext.strip().lower().replace(".", "")
    return SUPPORTED.get(e, "")

def _timestamp_name(prefix="file", ext="txt"):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.{ext}"

def make_plain_file(text: str, ext: str) -> io.BytesIO:
    bio = io.BytesIO()
    bio.write(text.encode("utf-8"))
    bio.seek(0)
    bio.name = _timestamp_name("text", ext)
    return bio

def make_pdf(text: str) -> io.BytesIO:
    # lightweight PDF using fpdf
    from fpdf import FPDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    # wrap to fit page
    for line in text.splitlines() or [" "]:
        for wrapped in textwrap.wrap(line, width=90) or [" "]:
            pdf.cell(0, 8, txt=wrapped, ln=True)
    out = io.BytesIO(pdf.output(dest="S").encode("latin1"))
    out.seek(0)
    out.name = _timestamp_name("text", "pdf")
    return out

def make_xlsx(text: str) -> io.BytesIO:
    # put the whole text in A1; also split lines into rows for convenience
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Text"
    lines = text.splitlines() or [""]
    for idx, line in enumerate(lines, start=1):
        ws.cell(row=idx, column=1, value=line)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    out.name = _timestamp_name("text", "xlsx")
    return out

def build_type_keyboard() -> InlineKeyboardMarkup:
    # Show a compact grid of file options
    order = ["txt", "py", "css", "js", "html", "json", "csv", "xml", "yaml", "md", "pdf", "xlsx"]
    buttons = []
    row = []
    for ext in order:
        row.append(InlineKeyboardButton(ext, callback_data=f"ext:{ext}"))
        if len(row) == 4:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    buttons.append([InlineKeyboardButton("⬅️ Start over", callback_data="startover")])
    return InlineKeyboardMarkup(buttons)

async def set_commands(app: Application):
    await app.bot.set_my_commands([
        BotCommand("start", "Open bot menu / welcome"),
        BotCommand("generate", "Start text → file generator"),
        BotCommand("help", "How to use")
    ])

# ---------- Handlers ----------
async def start(update: Update, context: CallbackContext) -> int:
    # Show a small reply keyboard like the screenshot vibe
    reply_kb = ReplyKeyboardMarkup(
        [[KeyboardButton("▶️ Start text to File Generate (/generate)")],
         [KeyboardButton("ℹ️ More menu"), KeyboardButton("❓ Help (/help)")]],
        resize_keyboard=True
    )
    msg = (
        "👋 **Welcome to Text ➜ Any File Bot!**\n\n"
        "এখানে যেকোনো টেক্সট থেকে এক ক্লিকে ফাইল বানাতে পারবে — txt, py, css, js, html, pdf, xls(xlsx), ইত্যাদি।\n\n"
        "কীভাবে ব্যবহার করবে:\n"
        "1) `/generate` প্রেস করো বা নিচের বাটন চাপো\n"
        "2) তোমার টেক্সট পাঠাও\n"
        "3) ফাইল টাইপ সিলেক্ট করো ✅\n\n"
        "Ready? শুরু করি! 🚀"
    )
    await update.effective_message.reply_text(msg, reply_markup=reply_kb, parse_mode="Markdown")
    return ConversationHandler.END

async def help_cmd(update: Update, context: CallbackContext):
    text = (
        "🛠 **Help**\n"
        "- `/generate` → টেক্সট পাঠাও, তারপর যেকোনো টাইপ বাছাই করো।\n"
        "- সাপোর্টেড: txt, py, css, js, html, json, csv, xml, yaml, md, pdf, xls/xlsx\n"
        "- অন্য মেসেজ দিলে বট বলবে: “Please /generate click.. Tmi boilo valo moto..” 😄"
    )
    await update.message.reply_text(text, parse_mode="Markdown")

async def more_menu(update: Update, context: CallbackContext):
    # simple alias for help
    await help_cmd(update, context)

async def generate(update: Update, context: CallbackContext) -> int:
    await update.effective_message.reply_text("📝 Please enter your text…")
    return ENTER_TEXT

async def got_text(update: Update, context: CallbackContext) -> int:
    context.user_data["text_to_convert"] = update.message.text or ""
    await update.message.reply_text(
        "✅ Got it!\n🔽 Choose your file type:",
        reply_markup=build_type_keyboard()
    )
    return CHOOSE_TYPE

async def choose_type_button(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "startover":
        await query.edit_message_text("📝 Please enter your text…")
        return ENTER_TEXT

    if not data.startswith("ext:"):
        await query.edit_message_text("Invalid choice. Try again:", reply_markup=build_type_keyboard())
        return CHOOSE_TYPE

    ext = data.split(":", 1)[1]  # normalized already
    text = context.user_data.get("text_to_convert", "")

    # Create file
    if ext in {"txt", "py", "css", "js", "html", "json", "csv", "xml", "yaml", "md"}:
        fileobj = make_plain_file(text, ext)
    elif ext == "pdf":
        fileobj = make_pdf(text)
    elif ext == "xlsx":
        fileobj = make_xlsx(text)
    else:
        await query.edit_message_text("Unsupported type. Pick again:", reply_markup=build_type_keyboard())
        return CHOOSE_TYPE

    await query.edit_message_text(f"📦 Generating `{ext}` file…", parse_mode="Markdown")
    await query.message.chat.send_document(document=fileobj, caption=f"Here is your `{ext}` file ✅", parse_mode="Markdown")
    return ConversationHandler.END

async def choose_type_text(update: Update, context: CallbackContext) -> int:
    # accept typed extensions during CHOOSE_TYPE
    ext = _normalize_ext(update.message.text)
    if not ext:
        await update.message.reply_text("❌ Unsupported type. Try again ↓", reply_markup=build_type_keyboard())
        return CHOOSE_TYPE
    # Simulate button flow
    fake_update = Update(update.update_id, callback_query=None)
    context.user_data.setdefault("text_to_convert", "")
    if ext == "xlsx":
        fileobj = make_xlsx(context.user_data["text_to_convert"])
    elif ext == "pdf":
        fileobj = make_pdf(context.user_data["text_to_convert"])
    else:
        fileobj = make_plain_file(context.user_data["text_to_convert"], ext)
    await update.message.reply_document(document=fileobj, caption=f"Here is your `{ext}` file ✅", parse_mode="Markdown")
    return ConversationHandler.END

async def fallback_msg(update: Update, context: CallbackContext):
    # Any other message outside conversations
    await update.message.reply_text("Please /generate click.. Tmi boilo valo moto.. 🙂")

async def error_handler(update: object, context: CallbackContext) -> None:
    try:
        raise context.error
    except Exception as e:
        if update and hasattr(update, "message") and update.message:
            await update.message.reply_text("⚠️ Unexpected error occurred. Try again.")
        print(f"ERROR: {e}")

def main():
    app = Application.builder().token(TOKEN).build()

    # Commands
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(MessageHandler(filters.Regex("^(ℹ️ More menu)$"), more_menu))

    # Generate flow
    conv = ConversationHandler(
        entry_points=[CommandHandler("generate", generate),
                      MessageHandler(filters.Regex("^▶️ Start text to File Generate"), generate)],
        states={
            ENTER_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_text)],
            CHOOSE_TYPE: [
                CallbackQueryHandler(choose_type_button),
                MessageHandler(filters.TEXT & ~filters.COMMAND, choose_type_text),
            ],
        },
        fallbacks=[CommandHandler("start", start), CommandHandler("help", help_cmd)],
        name="text_to_file_flow",
        persistent=False,
    )
    app.add_handler(conv)

    # Fallback for any other text
    app.add_handler(MessageHandler(filters.ALL & ~filters.COMMAND, fallback_msg))

    # Errors
    app.add_error_handler(error_handler)

    # Set commands so they appear in the Telegram menu (like your screenshot)
    app.post_init = set_commands

    print("Bot is running…")
    app.run_polling(close_loop=False)

if __name__ == "__main__":
    main()